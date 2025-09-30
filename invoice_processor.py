#!/usr/bin/env python3
"""
Invoice and PO Processing System with OCR
Processes invoices and purchase orders from PDFs and images,
extracts data, and saves to Excel with automatic linking.
"""

import os
import re
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pdfplumber
import pytesseract
from PIL import Image
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pdf2image import convert_from_path


class InvoiceProcessor:
    """Main class for processing invoices and POs"""
    
    def __init__(self, invoices_folder: str = "invoices", excel_file: str = "invoices.xlsx", log_file: str = "log.txt"):
        self.invoices_folder = Path(invoices_folder)
        self.excel_file = Path(excel_file)
        self.log_file = Path(log_file)
        
        # Supported file extensions
        self.pdf_extensions = ['.pdf']
        self.image_extensions = ['.jpg', '.jpeg', '.png', '.tiff', '.tif', '.bmp']
        
        # Setup logging
        self._setup_logging()
        
        # Initialize Excel file
        self._initialize_excel()
    
    def _setup_logging(self):
        """Setup logging configuration"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(self.log_file),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def _initialize_excel(self):
        """Initialize or load Excel file with required sheets"""
        if self.excel_file.exists():
            self.logger.info(f"Loading existing Excel file: {self.excel_file}")
            self.workbook = openpyxl.load_workbook(self.excel_file)
        else:
            self.logger.info(f"Creating new Excel file: {self.excel_file}")
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)  # Remove default sheet
            self._create_sheets()
        
        # Ensure both sheets exist
        if "PO_Details" not in self.workbook.sheetnames:
            self._create_po_sheet()
        if "Invoice_Details" not in self.workbook.sheetnames:
            self._create_invoice_sheet()
    
    def _create_sheets(self):
        """Create both required sheets"""
        self._create_po_sheet()
        self._create_invoice_sheet()
    
    def _create_po_sheet(self):
        """Create PO_Details sheet with headers"""
        ws = self.workbook.create_sheet("PO_Details")
        headers = ["Serial Number", "PO Number", "PO Date", "PO Amount", "Department"]
        ws.append(headers)
        self._style_header(ws)
        self.logger.info("Created PO_Details sheet")
    
    def _create_invoice_sheet(self):
        """Create Invoice_Details sheet with headers"""
        ws = self.workbook.create_sheet("Invoice_Details")
        headers = [
            "Serial Number", "Invoice Number", "Invoice Date", "PO Number",
            "PO Date", "Department", "GR ID", "GR Date", "Subtotal",
            "Tax 12%", "Grand Total", "Status"
        ]
        ws.append(headers)
        self._style_header(ws)
        self.logger.info("Created Invoice_Details sheet")
    
    def _style_header(self, worksheet):
        """Apply styling to header row"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _get_next_serial_number(self, sheet_name: str) -> int:
        """Get next serial number for a sheet"""
        ws = self.workbook[sheet_name]
        max_row = ws.max_row
        if max_row == 1:  # Only header exists
            return 1
        return max_row  # Next row number (1-indexed, minus header)
    
    def extract_text_from_pdf(self, pdf_path: Path) -> str:
        """Extract text from PDF using pdfplumber or OCR"""
        text = ""
        
        try:
            # Try text extraction first
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            
            # If no text found, use OCR
            if not text.strip():
                self.logger.info(f"No text layer found in {pdf_path.name}, using OCR")
                text = self._ocr_pdf(pdf_path)
            else:
                self.logger.info(f"Extracted text from {pdf_path.name} using pdfplumber")
        
        except Exception as e:
            self.logger.error(f"Error extracting text from {pdf_path.name}: {str(e)}")
            self._log_error(pdf_path.name, str(e))
        
        return text
    
    def _ocr_pdf(self, pdf_path: Path) -> str:
        """Perform OCR on PDF by converting to images"""
        text = ""
        try:
            images = convert_from_path(pdf_path)
            for i, image in enumerate(images):
                self.logger.info(f"OCR processing page {i+1} of {pdf_path.name}")
                page_text = pytesseract.image_to_string(image)
                text += page_text + "\n"
        except Exception as e:
            self.logger.error(f"OCR error on {pdf_path.name}: {str(e)}")
            self._log_error(pdf_path.name, f"OCR error: {str(e)}")
        
        return text
    
    def extract_text_from_image(self, image_path: Path) -> str:
        """Extract text from image using OCR"""
        text = ""
        try:
            image = Image.open(image_path)
            text = pytesseract.image_to_string(image)
            self.logger.info(f"OCR extracted text from {image_path.name}")
        except Exception as e:
            self.logger.error(f"Error processing image {image_path.name}: {str(e)}")
            self._log_error(image_path.name, str(e))
        
        return text
    
    def _parse_table_format(self, text: str) -> Dict:
        """Parse table format where headers and values are on separate lines"""
        data = {}
        
        # Look for the table pattern: PO NO | PO DATE | GRNO | GR DATE
        # followed by values on the next line(s)
        # Multiple patterns to handle various formats
        table_patterns = [
            # Pattern 1: Standard format with pipes
            r'PO\s*NO[\s|]*PO\s*DATE[\s|]*GR\s*NO[\s|]*GR\s*DATE.*?\n\s*(\d+)\s+([\d-]+[-/]\w+[-/][\d]+)[\s|]*(\d+)\s+([\d-]+[-/]\w+[-/][\d]+)',
            # Pattern 2: With extra pipes between GR NO and GR DATE (like 10028)
            r'PO\s*NO[\s|]*PO\s*DATE[\s|]*GR\s*NO[\s|]*GR\s*DATE.*?\n\s*(\d+)[\s|]+([\d-]+[-/]\w+[-/][\d]+)[\s|]+([\d]+)[\s|]+([\d-]+[-/]\w+[-/][\d]+)',
            # Pattern 3: Compact format with specific spacing
            r'PO\s*NO[\s|]+PO\s*DATE[\s|]+GR\s*NO[\s|]+GR\s*DATE[\s\S]*?(\d{10})[\s|]+([\d-]+[-/]\w+[-/][\d]+)[\s|]+([\d]{7})[\s|]+([\d-]+[-/]\w+[-/][\d]+)',
            # Pattern 4: PODATE without space (like 1523)
            r'PO\s*NO[\s|]*PODATE[\s|]*GR\s*NO[\s\S]*?(\d{10})\s+([\d-]+[-/]\w+[-/][\d]+)[\s|]+([\d]{7})\s+([\d-]+[-/]\w+[-/][\d]+)',
        ]
        
        for pattern in table_patterns:
            table_match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
            if table_match:
                data['po_number'] = table_match.group(1).strip()
                data['po_date'] = self._parse_date(table_match.group(2).strip())
                data['gr_id'] = table_match.group(3).strip()
                data['gr_date'] = self._parse_date(table_match.group(4).strip())
                self.logger.info(f"Extracted from table: PO={data['po_number']}, GR={data['gr_id']}")
                break
        
        return data
    
    def parse_invoice_data(self, text: str) -> Optional[Dict]:
        """Parse invoice data from extracted text"""
        data = {}
        
        try:
            # Log extracted text for debugging (first 500 chars)
            self.logger.debug(f"Extracted text preview: {text[:500]}...")
            
            # First try to parse table format (common in these invoices)
            table_data = self._parse_table_format(text)
            data.update(table_data)
            
            # Invoice Number - capture FULL number including prefix (A10001, 10001, etc.)
            invoice_patterns = [
                r'Invoice\s*No[:\s.]*[:\s]*([A-Z]?\d+)',  # Captures A10001 or 10001 fully
                r'Invoice\s*#[:\s]*([A-Z]?\d+)',
                r'Inv[\s.]*No[:\s.]*[:\s]*([A-Z]?\d+)',
                r'Invoice\s*Number[:\s]*([A-Z]?\d+)',
                r'Invoice\s*No[:\s.]*\s*$.*?^.*?([A-Z]?\d{4,})',  # Invoice No: on one line, number on next
            ]
            for pattern in invoice_patterns:
                invoice_match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                if invoice_match:
                    data['invoice_number'] = invoice_match.group(1)
                    break
            
            # Invoice Date - more flexible patterns
            date_patterns = [
                r'Invoice\s*Date[:\s.]*[:\s]*(\d{1,2}[-/]\w+[-/]\d{2,4})',
                r'Invoice\s*Date[:\s.]*[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
                r'[Ii]nvoice[\s\S]{0,30}Date[:\s]*(\d{1,2}[-/]\w+[-/]\d{2,4})',  # Flexible spacing
                r'[Ii]woie[\s\S]{0,20}Date[:\s]*(\d{1,2}[-/]\w+[-/]\d{2,4})',  # OCR error: Invoice -> Iwoie
                r'iavoie[\s\S]{0,20}Date[:\s]*(\d{1,2}[-/]\w+[-/]\d{2,4})',  # OCR error: Invoice -> iavoie
                r'[Ii]nvoice\s*No[:\s]*\d+[\s\S]{0,100}?(\d{1,2}[-/]\w+[-/]\d{2,4})',  # Date near Invoice No
            ]
            for pattern in date_patterns:
                date_match = re.search(pattern, text, re.IGNORECASE)
                if date_match:
                    data['invoice_date'] = self._parse_date(date_match.group(1))
                    break
            
            # PO Number - only if not found in table
            if 'po_number' not in data:
                po_patterns = [
                    r'PO\s*NO[:\s.]*[:\s]*(\d+)',
                    r'PO\s*Number[:\s]*(\d+)',
                    r'P\.?O\.?[:\s]*(\d+)',
                    r'Purchase\s*Order[:\s]*(\d+)'
                ]
                for pattern in po_patterns:
                    po_match = re.search(pattern, text, re.IGNORECASE)
                    if po_match:
                        data['po_number'] = po_match.group(1)
                        break
            
            # PO Date - only if not found in table
            if 'po_date' not in data:
                po_date_patterns = [
                    r'PO\s*DATE[:\s.]*[:\s]*(\d{1,2}[-/]\w+[-/]\d{2,4})',
                    r'PO\s*DATE[:\s.]*[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
                    r'P\.?O\.?\s*Date[:\s]*(\d{1,2}[-/]\w+[-/]\d{2,4})'
                ]
                for pattern in po_date_patterns:
                    po_date_match = re.search(pattern, text, re.IGNORECASE)
                    if po_date_match:
                        data['po_date'] = self._parse_date(po_date_match.group(1))
                        break
            
            # GR ID - only if not found in table
            if 'gr_id' not in data:
                gr_patterns = [
                    r'GR\s*NO[:\s.]*[:\s]*(\d+)',
                    r'GR\s*Number[:\s]*(\d+)',
                    r'G\.?R\.?[:\s]*(\d+)',
                    r'Goods\s*Receipt[:\s]*(\d+)'
                ]
                for pattern in gr_patterns:
                    gr_match = re.search(pattern, text, re.IGNORECASE)
                    if gr_match:
                        data['gr_id'] = gr_match.group(1)
                        break
            
            # GR Date - only if not found in table
            if 'gr_date' not in data:
                gr_date_patterns = [
                    r'GR\s*DATE[:\s.]*[:\s]*(\d{1,2}[-/]\w+[-/]\d{2,4})',
                    r'GR\s*DATE[:\s.]*[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
                    r'G\.?R\.?\s*Date[:\s]*(\d{1,2}[-/]\w+[-/]\d{2,4})'
                ]
                for pattern in gr_date_patterns:
                    gr_date_match = re.search(pattern, text, re.IGNORECASE)
                    if gr_date_match:
                        data['gr_date'] = self._parse_date(gr_date_match.group(1))
                        break
            
            # Subtotal/Total - more flexible patterns
            total_patterns = [
                r'(?:^|\n)\s*TOTAL[:\s]+([0-9,]+\.?\d*)',
                r'Sub\s*Total[:\s]+([0-9,]+\.?\d*)',
                r'Amount[:\s]+([0-9,]+\.?\d*)'
            ]
            for pattern in total_patterns:
                total_match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                if total_match:
                    data['subtotal'] = self._parse_amount(total_match.group(1))
                    break
            
            # Tax (KPRA or other) - more flexible patterns
            tax_patterns = [
                r'(?:KPRA|Tax)\s*\d+%[:\s]+([0-9,]+\.?\d*)',
                r'(?:KPRA|Tax)[:\s]+([0-9,]+\.?\d*)',
                r'VAT\s*\d+%[:\s]+([0-9,]+\.?\d*)'
            ]
            for pattern in tax_patterns:
                tax_match = re.search(pattern, text, re.IGNORECASE)
                if tax_match:
                    data['tax'] = self._parse_amount(tax_match.group(1))
                    break
            
            # Grand Total - more flexible patterns
            grand_total_patterns = [
                r'GRAND\s*TOTAL[:\s]+([0-9,]+\.?\d*)',
                r'Total\s*Amount[:\s]+([0-9,]+\.?\d*)',
                r'Net\s*Total[:\s]+([0-9,]+\.?\d*)'
            ]
            for pattern in grand_total_patterns:
                grand_total_match = re.search(pattern, text, re.IGNORECASE)
                if grand_total_match:
                    data['grand_total'] = self._parse_amount(grand_total_match.group(1))
                    break
            
            # Calculate tax if not found (12% of subtotal)
            if 'subtotal' in data and 'tax' not in data:
                data['tax'] = round(data['subtotal'] * 0.12, 2)
            
            # Calculate grand total if not found
            if 'subtotal' in data and 'grand_total' not in data:
                tax = data.get('tax', 0)
                data['grand_total'] = data['subtotal'] + tax
            
            # Status - default to UnPaid
            data['status'] = 'UnPaid'
            
            # Log what was found
            self.logger.info(f"Parsed data: Invoice={data.get('invoice_number', 'N/A')}, "
                           f"Date={data.get('invoice_date', 'N/A')}, "
                           f"PO={data.get('po_number', 'N/A')}, "
                           f"GR={data.get('gr_id', 'N/A')}")
            
            return data if data else None
        
        except Exception as e:
            self.logger.error(f"Error parsing invoice data: {str(e)}")
            return None
    
    def parse_po_data(self, text: str) -> Optional[Dict]:
        """Parse PO data from extracted text"""
        data = {}
        
        try:
            # PO Number
            po_match = re.search(r'PO\s*(?:Number|NO)[:\s]+(\d+)', text, re.IGNORECASE)
            if po_match:
                data['po_number'] = po_match.group(1)
            
            # PO Date
            date_match = re.search(r'PO\s*DATE[:\s]+(\d{1,2}[-/]\w+[-/]\d{2,4})', text, re.IGNORECASE)
            if date_match:
                data['po_date'] = self._parse_date(date_match.group(1))
            
            # PO Amount
            amount_match = re.search(r'(?:Amount|Total)[:\s]+([0-9,]+\.?\d*)', text, re.IGNORECASE)
            if amount_match:
                data['po_amount'] = self._parse_amount(amount_match.group(1))
            
            # Department (if mentioned)
            dept_match = re.search(r'Department[:\s]+([A-Za-z\s]+)', text, re.IGNORECASE)
            if dept_match:
                data['department'] = dept_match.group(1).strip()
            else:
                data['department'] = 'N/A'
            
            return data if 'po_number' in data else None
        
        except Exception as e:
            self.logger.error(f"Error parsing PO data: {str(e)}")
            return None
    
    def _parse_date(self, date_str: str) -> str:
        """Parse date string to standard format"""
        try:
            # Try different date formats
            formats = [
                '%d-%b-%y', '%d-%b-%Y', '%d/%b/%y', '%d/%b/%Y',
                '%d-%m-%y', '%d-%m-%Y', '%d/%m/%y', '%d/%m/%Y'
            ]
            
            for fmt in formats:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.strftime('%d-%b-%Y')
                except ValueError:
                    continue
            
            # If no format matches, return as is
            return date_str
        except:
            return date_str
    
    def _parse_amount(self, amount_str: str) -> float:
        """Parse amount string to float"""
        try:
            # Remove commas and convert to float
            return float(amount_str.replace(',', ''))
        except:
            return 0.0
    
    def _lookup_po_details(self, po_number: str) -> Tuple[Optional[str], Optional[str]]:
        """Lookup PO Date and Department from PO_Details sheet"""
        ws = self.workbook["PO_Details"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == po_number:  # PO Number is in column 1 (index 1)
                return row[2], row[4]  # PO Date (index 2), Department (index 4)
        
        return None, None
    
    def add_po_record(self, po_data: Dict):
        """Add PO record to PO_Details sheet"""
        ws = self.workbook["PO_Details"]
        serial = self._get_next_serial_number("PO_Details")
        
        row = [
            serial,
            po_data.get('po_number', ''),
            po_data.get('po_date', ''),
            po_data.get('po_amount', 0),
            po_data.get('department', 'N/A')
        ]
        
        ws.append(row)
        self.logger.info(f"Added PO record: {po_data.get('po_number')}")
    
    def _invoice_exists(self, invoice_number: str) -> bool:
        """Check if invoice number already exists in Invoice_Details sheet"""
        ws = self.workbook["Invoice_Details"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == invoice_number:  # Invoice Number is in column 1 (index 1)
                return True
        
        return False
    
    def add_invoice_record(self, invoice_data: Dict):
        """Add invoice record to Invoice_Details sheet with PO linking"""
        invoice_number = invoice_data.get('invoice_number', '')
        
        # Check if invoice already exists
        if invoice_number and self._invoice_exists(invoice_number):
            self.logger.info(f"Invoice {invoice_number} already exists - skipping")
            return
        
        ws = self.workbook["Invoice_Details"]
        serial = self._get_next_serial_number("Invoice_Details")
        
        # Lookup PO details if PO number exists
        po_number = invoice_data.get('po_number', '')
        po_date_from_sheet, department_from_sheet = None, None
        
        if po_number:
            po_date_from_sheet, department_from_sheet = self._lookup_po_details(po_number)
            
            if not po_date_from_sheet:
                # PO not found - auto-create it from invoice data
                self.logger.info(f"PO Number {po_number} not found in PO_Details - auto-creating from invoice data")
                po_data = {
                    'po_number': po_number,
                    'po_date': invoice_data.get('po_date', ''),
                    'po_amount': 0,  # Amount not available from invoice
                    'department': 'N/A'  # Department not available from invoice
                }
                self.add_po_record(po_data)
                # Now lookup again to get the values
                po_date_from_sheet, department_from_sheet = self._lookup_po_details(po_number)
        
        # Use PO date from sheet if available, otherwise from invoice
        po_date = po_date_from_sheet or invoice_data.get('po_date', '')
        department = department_from_sheet or 'N/A'
        
        row = [
            serial,
            invoice_data.get('invoice_number', ''),
            invoice_data.get('invoice_date', ''),
            po_number,
            po_date,
            department,
            invoice_data.get('gr_id', ''),
            invoice_data.get('gr_date', ''),
            invoice_data.get('subtotal', 0),
            invoice_data.get('tax', 0),
            invoice_data.get('grand_total', 0),
            invoice_data.get('status', 'UnPaid')
        ]
        
        ws.append(row)
        self.logger.info(f"Added invoice record: {invoice_data.get('invoice_number')}")
    
    def _log_error(self, filename: str, error: str):
        """Log error to log file"""
        with open(self.log_file, 'a') as f:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            f.write(f"[{timestamp}] {filename}: {error}\n")
    
    def process_file(self, file_path: Path):
        """Process a single file (PDF or image)"""
        self.logger.info(f"Processing file: {file_path.name}")
        
        try:
            # Extract text
            if file_path.suffix.lower() in self.pdf_extensions:
                text = self.extract_text_from_pdf(file_path)
            elif file_path.suffix.lower() in self.image_extensions:
                text = self.extract_text_from_image(file_path)
            else:
                self.logger.warning(f"Unsupported file format: {file_path.name}")
                return
            
            if not text.strip():
                self.logger.warning(f"No text extracted from {file_path.name}")
                self._log_error(file_path.name, "No text could be extracted")
                return
            
            # Save extracted text for debugging
            debug_folder = Path("debug_ocr_text")
            debug_folder.mkdir(exist_ok=True)
            debug_file = debug_folder / f"{file_path.stem}_extracted.txt"
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write(text)
            self.logger.info(f"Saved extracted text to: {debug_file}")
            
            # Determine if it's an invoice or PO based on content
            is_invoice = 'invoice' in text.lower() and 'invoice no' in text.lower()
            is_po = 'purchase order' in text.lower() or ('po no' in text.lower() and 'invoice' not in text.lower())
            
            if is_invoice:
                invoice_data = self.parse_invoice_data(text)
                if invoice_data:
                    self.add_invoice_record(invoice_data)
                else:
                    self.logger.warning(f"Could not parse invoice data from {file_path.name}")
                    self._log_error(file_path.name, "Failed to parse invoice data")
            
            elif is_po:
                po_data = self.parse_po_data(text)
                if po_data:
                    self.add_po_record(po_data)
                else:
                    self.logger.warning(f"Could not parse PO data from {file_path.name}")
                    self._log_error(file_path.name, "Failed to parse PO data")
            
            else:
                self.logger.warning(f"Could not determine document type for {file_path.name}")
                self._log_error(file_path.name, "Unknown document type")
        
        except Exception as e:
            self.logger.error(f"Error processing {file_path.name}: {str(e)}")
            self._log_error(file_path.name, str(e))
    
    def process_all_files(self):
        """Process all files in the invoices folder"""
        if not self.invoices_folder.exists():
            self.logger.error(f"Invoices folder not found: {self.invoices_folder}")
            self.invoices_folder.mkdir(parents=True, exist_ok=True)
            self.logger.info(f"Created invoices folder: {self.invoices_folder}")
            return
        
        # Get all supported files
        files = []
        for ext in self.pdf_extensions + self.image_extensions:
            files.extend(self.invoices_folder.glob(f"*{ext}"))
        
        if not files:
            self.logger.warning(f"No files found in {self.invoices_folder}")
            return
        
        self.logger.info(f"Found {len(files)} files to process")
        
        for file_path in files:
            self.process_file(file_path)
        
        # Save Excel file
        self.save_excel()
    
    def save_excel(self):
        """Save the Excel workbook"""
        try:
            self.workbook.save(self.excel_file)
            self.logger.info(f"Excel file saved: {self.excel_file}")
        except Exception as e:
            self.logger.error(f"Error saving Excel file: {str(e)}")
            self._log_error("Excel Save", str(e))


def main():
    """Main entry point"""
    print("=" * 60)
    print("Invoice and PO Processing System")
    print("=" * 60)
    
    # Initialize processor
    processor = InvoiceProcessor(
        invoices_folder="invoices",
        excel_file="invoices.xlsx",
        log_file="log.txt"
    )
    
    # Process all files
    processor.process_all_files()
    
    print("=" * 60)
    print("Processing complete!")
    print(f"Results saved to: {processor.excel_file}")
    print(f"Logs saved to: {processor.log_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
